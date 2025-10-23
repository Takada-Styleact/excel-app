"""
Excel読み書きモジュール

このモジュールは、ExcelファイルとCSVファイルの読み込みと書き込みを担当します。
- 備考行（1行目）を保持しながら読み込み（Excelの場合）
- CSVファイルの読み込み対応（複数エンコーディング自動検出）
- 3シート構成のExcelファイルを作成
"""

import pandas as pd
import openpyxl
from io import BytesIO, StringIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def detect_csv_encoding(file):
    """
    CSVファイルのエンコーディングを自動検出

    Args:
        file: ファイルオブジェクト

    Returns:
        str: 検出されたエンコーディング名
    """
    # よく使われるエンコーディングを試す
    encodings = ['utf-8-sig', 'utf-8', 'shift_jis', 'cp932', 'iso-2022-jp', 'euc-jp']

    if hasattr(file, 'seek'):
        file.seek(0)

    # ファイルの先頭を読み取る
    sample = file.read(10000) if hasattr(file, 'read') else b''

    if hasattr(file, 'seek'):
        file.seek(0)

    # 各エンコーディングを試す
    for encoding in encodings:
        try:
            if isinstance(sample, bytes):
                sample.decode(encoding)
            return encoding
        except (UnicodeDecodeError, LookupError):
            continue

    # デフォルトはUTF-8
    return 'utf-8'


def read_excel_with_comment(file):
    """
    ExcelファイルまたはCSVファイルを読み込み、備考行とデータを返す

    Excel: 1行目が備考行の場合と、1行目がヘッダー行の場合の両方に対応
    CSV: 1行目がヘッダー行として読み込み（備考行は空）

    Args:
        file: Streamlitのアップロードファイル or ファイルパス

    Returns:
        tuple: (DataFrame, 備考行の文字列)

    Raises:
        ValueError: ファイル読み込みエラー
    """
    try:
        # ファイル名から拡張子を判定
        file_name = file.name if hasattr(file, 'name') else str(file)
        is_csv = file_name.lower().endswith('.csv')

        if is_csv:
            # CSVファイルの処理
            # エンコーディングを自動検出
            encoding = detect_csv_encoding(file)

            # ファイルポインタをリセット
            if hasattr(file, 'seek'):
                file.seek(0)

            # まず2行目をヘッダーとして読み込み（備考行がある場合を想定）
            df = pd.read_csv(file, header=1, encoding=encoding)

            # ファイルポインタをリセット
            if hasattr(file, 'seek'):
                file.seek(0)

            # 1行目の内容を取得（備考行の可能性）
            first_line = file.readline()
            if isinstance(first_line, bytes):
                first_line = first_line.decode(encoding)
            comment_row = first_line.strip()

            # ファイルポインタをリセット
            if hasattr(file, 'seek'):
                file.seek(0)

            # 必須カラムのチェック
            required_columns = ['stationid', 'railroad']
            df_columns_lower = [col.lower() if isinstance(col, str) else col for col in df.columns]
            has_required_columns = all(
                req_col.lower() in df_columns_lower for req_col in required_columns
            )

            # 必須カラムがない場合、1行目をヘッダーとして再試行
            if not has_required_columns:
                if hasattr(file, 'seek'):
                    file.seek(0)
                df_alt = pd.read_csv(file, header=0, encoding=encoding)
                df_alt_columns_lower = [col.lower() if isinstance(col, str) else col for col in df_alt.columns]
                has_required_columns_alt = all(
                    req_col.lower() in df_alt_columns_lower for req_col in required_columns
                )

                if has_required_columns_alt:
                    df = df_alt
                    comment_row = ""  # 備考行なし

            return df, comment_row

        else:
            # Excelファイルの処理（既存のロジック）
            # openpyxlで1行目を取得
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            first_row_value = ws[1][0].value if ws[1][0].value else ""

            # ファイルポインタをリセット（ファイルオブジェクトの場合のみ）
            if hasattr(file, 'seek'):
                file.seek(0)

            # まず2行目をヘッダーとして読み込み
            df = pd.read_excel(file, header=1)

            # 必須カラムのチェック
            required_columns = ['stationid', 'railroad']
            df_columns_lower = [col.lower() if isinstance(col, str) else col for col in df.columns]
            has_required_columns = all(
                req_col.lower() in df_columns_lower for req_col in required_columns
            )

            # 必須カラムがない場合、1行目をヘッダーとして再試行
            if not has_required_columns:
                # ファイルポインタをリセット
                if hasattr(file, 'seek'):
                    file.seek(0)

                # 1行目をヘッダーとして読み込み
                df_alt = pd.read_excel(file, header=0)
                df_alt_columns_lower = [col.lower() if isinstance(col, str) else col for col in df_alt.columns]
                has_required_columns_alt = all(
                    req_col.lower() in df_alt_columns_lower for req_col in required_columns
                )

                # 1行目にヘッダーがある場合
                if has_required_columns_alt:
                    df = df_alt
                    comment_row = ""  # 備考行なし
                else:
                    # どちらにもない場合は2行目をヘッダーとして使用（元の動作）
                    comment_row = str(first_row_value) if first_row_value else ""
            else:
                # 2行目がヘッダーの場合は1行目を備考行とする
                comment_row = str(first_row_value) if first_row_value else ""

            return df, comment_row

    except Exception as e:
        raise ValueError(f"ファイルの読み込みエラー: {str(e)}")


def write_excel_with_sheets(
    sheet1_df, sheet1_comment,
    sheet2_df, sheet2_comment,
    sheet3_df, sheet3_comment,
    sheet4_df=None, sheet4_comment=""
):
    """
    3シートまたは4シートのExcelファイルを作成

    Args:
        sheet1_df: シート1のDataFrame
        sheet1_comment: シート1の備考行（文字列 or 辞書）
        sheet2_df: シート2のDataFrame
        sheet2_comment: シート2の備考行（文字列 or 辞書）
        sheet3_df: シート3のDataFrame
        sheet3_comment: シート3の備考行（文字列 or 辞書）
        sheet4_df: シート4のDataFrame（オプション）
        sheet4_comment: シート4の備考行（オプション、文字列 or 辞書）

    Returns:
        BytesIO: Excelファイルのバイナリ
    """
    output = BytesIO()

    # まずpandasでベースを作成（1行目に備考行用スペースを空けてデータを書き込む）
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet1_df.to_excel(writer, sheet_name='前回データ', index=False, startrow=1)
        sheet2_df.to_excel(writer, sheet_name='今回データ', index=False, startrow=1)
        sheet3_df.to_excel(writer, sheet_name='比較データ', index=False, startrow=1)
        if sheet4_df is not None:
            sheet4_df.to_excel(writer, sheet_name='異常値シート', index=False, startrow=1)

    # openpyxlで備考行を1行目に追加
    output.seek(0)
    wb = load_workbook(output)

    # 各シートの1行目に備考行を書き込む
    sheets_to_process = [
        ('前回データ', sheet1_comment),
        ('今回データ', sheet2_comment),
        ('比較データ', sheet3_comment)
    ]
    if sheet4_df is not None:
        sheets_to_process.append(('異常値シート', sheet4_comment))

    for sheet_name, comment in sheets_to_process:
        ws = wb[sheet_name]

        # commentが辞書の場合は各セルに個別に設定
        if isinstance(comment, dict):
            for cell_position, text in comment.items():
                ws[cell_position] = text
        else:
            # 文字列の場合はA1に設定（従来通り）
            ws['A1'] = comment

    # セルの背景色を設定
    # 黄色: #FFF266
    yellow_fill = PatternFill(start_color='FFF266', end_color='FFF266', fill_type='solid')
    # 橙色: #FFD2B3
    orange_fill = PatternFill(start_color='FFD2B3', end_color='FFD2B3', fill_type='solid')

    # 前回データシートの色付け（2行目 = ヘッダー行）
    ws_previous = wb['前回データ']
    ws_previous['J2'].fill = yellow_fill  # J列を黄色
    ws_previous['L2'].fill = yellow_fill  # L列を黄色
    ws_previous['M2'].fill = orange_fill  # M列を橙色

    # 今回データシートの色付け（2行目 = ヘッダー行）
    ws_current = wb['今回データ']
    ws_current['J2'].fill = yellow_fill  # J列を黄色
    ws_current['L2'].fill = yellow_fill  # L列を黄色
    ws_current['M2'].fill = orange_fill  # M列を橙色

    # 比較データシートの色付け（2行目 = ヘッダー行）
    ws_comparison = wb['比較データ']
    ws_comparison['F2'].fill = yellow_fill  # F列を黄色
    ws_comparison['G2'].fill = yellow_fill  # G列を黄色
    ws_comparison['H2'].fill = orange_fill  # H列を橙色
    ws_comparison['I2'].fill = orange_fill  # I列を橙色

    # 異常値シートの色付け（2行目 = ヘッダー行）
    if sheet4_df is not None and '異常値シート' in wb.sheetnames:
        ws_abnormal = wb['異常値シート']
        ws_abnormal['F2'].fill = yellow_fill  # F列を黄色
        ws_abnormal['G2'].fill = yellow_fill  # G列を黄色
        ws_abnormal['H2'].fill = orange_fill  # H列を橙色
        ws_abnormal['I2'].fill = orange_fill  # I列を橙色

    # BytesIOに保存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
